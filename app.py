import io
import re
import zipfile
from copy import copy

import streamlit as st
from openpyxl import load_workbook


st.set_page_config(page_title="Headcount Splitter", page_icon="📊", layout="wide")
st.title("Headcount Splitter")
st.caption("Upload a current headcount workbook and create one Excel file per department.")


def safe_filename(value: str) -> str:
    value = str(value).strip()
    value = re.sub(r'[\\/:*?"<>|]+', '_', value)
    value = re.sub(r'\s+', ' ', value)
    return value[:120] or "Unassigned"


def copy_cell(source, target):
    target.value = source.value
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
    if source.hyperlink:
        target._hyperlink = copy(source.hyperlink)
    if source.comment:
        target.comment = copy(source.comment)


def build_department_workbook(source_ws, header_row, dept_col_idx, dept_value):
    from openpyxl import Workbook

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Headcount"

    # Preserve basic sheet settings.
    out_ws.freeze_panes = source_ws.freeze_panes
    out_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines

    for col_letter, dimension in source_ws.column_dimensions.items():
        out_ws.column_dimensions[col_letter].width = dimension.width
        out_ws.column_dimensions[col_letter].hidden = dimension.hidden

    # Copy all rows through the selected header row.
    target_row = 1
    for row_idx in range(1, header_row + 1):
        out_ws.row_dimensions[target_row].height = source_ws.row_dimensions[row_idx].height
        for col_idx in range(1, source_ws.max_column + 1):
            copy_cell(source_ws.cell(row_idx, col_idx), out_ws.cell(target_row, col_idx))
        target_row += 1

    # Copy matching employee rows.
    for row_idx in range(header_row + 1, source_ws.max_row + 1):
        current_value = source_ws.cell(row_idx, dept_col_idx).value
        normalized = "Unassigned" if current_value is None or str(current_value).strip() == "" else str(current_value).strip()
        if normalized == dept_value:
            out_ws.row_dimensions[target_row].height = source_ws.row_dimensions[row_idx].height
            for col_idx in range(1, source_ws.max_column + 1):
                copy_cell(source_ws.cell(row_idx, col_idx), out_ws.cell(target_row, col_idx))
            target_row += 1

    out_ws.auto_filter.ref = f"A{header_row}:{out_ws.cell(out_ws.max_row, out_ws.max_column).coordinate}"

    buffer = io.BytesIO()
    out_wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), target_row - header_row - 1


uploaded_file = st.file_uploader("Upload current headcount file", type=["xlsx"])

if uploaded_file:
    file_bytes = uploaded_file.getvalue()
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
    except Exception as exc:
        st.error(f"Could not open this workbook: {exc}")
        st.stop()

    sheet_name = st.selectbox("Worksheet containing headcount data", wb.sheetnames)
    ws = wb[sheet_name]

    header_row = st.number_input(
        "Header row",
        min_value=1,
        max_value=max(ws.max_row, 1),
        value=1,
        step=1,
        help="The row containing column names such as Employee ID, Name and Department.",
    )
    header_row = int(header_row)

    headers = []
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col_idx).value
        headers.append(str(value).strip() if value is not None else f"Column {col_idx}")

    guessed_index = next(
        (i for i, h in enumerate(headers) if h.lower() in {"department", "dept", "business unit", "business_unit"}),
        0,
    )
    department_header = st.selectbox("Department column", headers, index=guessed_index)
    dept_col_idx = headers.index(department_header) + 1

    departments = []
    counts = {}
    for row_idx in range(header_row + 1, ws.max_row + 1):
        value = ws.cell(row_idx, dept_col_idx).value
        dept = "Unassigned" if value is None or str(value).strip() == "" else str(value).strip()
        if dept not in counts:
            departments.append(dept)
            counts[dept] = 0
        counts[dept] += 1

    if not departments:
        st.warning("No data rows were found beneath the selected header row.")
        st.stop()

    total_people = sum(counts.values())
    col1, col2 = st.columns(2)
    col1.metric("Rows detected", total_people)
    col2.metric("Departments detected", len(departments))

    st.subheader("Preview")
    st.dataframe(
        [{"Department": dept, "Headcount rows": counts[dept]} for dept in departments],
        use_container_width=True,
        hide_index=True,
    )

    selected_departments = st.multiselect(
        "Departments to generate",
        options=departments,
        default=departments,
    )

    if st.button("Generate department files", type="primary", disabled=not selected_departments):
        zip_buffer = io.BytesIO()
        manifest = []
        with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            used_names = set()
            for dept in selected_departments:
                workbook_bytes, row_count = build_department_workbook(
                    ws, header_row, dept_col_idx, dept
                )
                base = safe_filename(dept)
                filename = f"{base}.xlsx"
                suffix = 2
                while filename.lower() in used_names:
                    filename = f"{base}_{suffix}.xlsx"
                    suffix += 1
                used_names.add(filename.lower())
                zf.writestr(filename, workbook_bytes)
                manifest.append((filename, row_count))

        zip_buffer.seek(0)
        st.success(f"Created {len(manifest)} department files.")
        st.download_button(
            "Download all department files (.zip)",
            data=zip_buffer.getvalue(),
            file_name="headcount_by_department.zip",
            mime="application/zip",
        )
